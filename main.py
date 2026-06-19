import cv2
import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
from PIL import Image, ImageTk
from deepface import DeepFace

# --- Setup Modern Theme ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class FaceAttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pro Face Attendance Dashboard")
        self.root.geometry("1100x750")
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Admin Security Settings
        self.is_admin_authenticated = False
        self.admin_password = "admin123"  # Change this to your preferred password

        # Database and Excel Setup
        self.db_folder = "database"
        self.excel_file = "attendance.xlsx"
        self.setup_files()

        # Registration Variables
        self.reg_name = ""
        self.reg_roll = ""

        # Camera Variables
        self.cap = None
        self.is_camera_running = False
        self.current_frame = None

        # Build GUI
        self.build_ui()
        
        # Auto-start camera
        self.start_camera()

    def setup_files(self):
        if not os.path.exists(self.db_folder):
            os.makedirs(self.db_folder)
        
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Roll No", "Date", "Time"])
            wb.save(self.excel_file)

    def build_ui(self):
        # Configure main grid layout
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

        # ==================== SIDEBAR ====================
        self.sidebar_frame = ctk.CTkFrame(self.root, width=220, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1) 

        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="AI Attendance", font=ctk.CTkFont(size=24, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(30, 40))

        self.btn_register = ctk.CTkButton(self.sidebar_frame, text="Register New User", 
                                          command=self.register_ui, height=45, font=ctk.CTkFont(size=14),
                                          fg_color="#2980b9", hover_color="#3498db")
        self.btn_register.grid(row=1, column=0, padx=20, pady=10)

        self.btn_attendance = ctk.CTkButton(self.sidebar_frame, text="Mark Attendance", 
                                            command=self.attendance_ui, height=45, font=ctk.CTkFont(size=14),
                                            fg_color="#2ecc71", hover_color="#27ae60")
        self.btn_attendance.grid(row=2, column=0, padx=20, pady=10)

        self.btn_exit = ctk.CTkButton(self.sidebar_frame, text="Exit System", 
                                      command=self.on_closing, height=45, font=ctk.CTkFont(size=14),
                                      fg_color="#e74c3c", hover_color="#c0392b")
        self.btn_exit.grid(row=5, column=0, padx=20, pady=(10, 30))

        # ==================== MAIN WORKSPACE ====================
        self.main_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, padx=30, pady=30, sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(2, weight=1) 

        # 1. Main Title
        self.title_label = ctk.CTkLabel(self.main_frame, text="System Dashboard", font=ctk.CTkFont(size=28, weight="bold"))
        self.title_label.grid(row=0, column=0, sticky="w", pady=(0, 20))

        # 2A. Admin Login Container (Inline)
        self.admin_login_frame = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="#2b2b2b")
        self.admin_login_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.admin_login_frame, text="Admin Security Check", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, pady=(15, 5))
        
        self.entry_admin_pass = ctk.CTkEntry(self.admin_login_frame, placeholder_text="Enter Admin Password...", show="*", height=40, font=ctk.CTkFont(size=14), width=250)
        self.entry_admin_pass.grid(row=1, column=0, pady=(5, 5))
        self.entry_admin_pass.bind("<Return>", lambda event: self.verify_admin()) 
        
        self.lbl_login_error = ctk.CTkLabel(self.admin_login_frame, text="", text_color="#e74c3c", font=ctk.CTkFont(size=12))
        self.lbl_login_error.grid(row=2, column=0, pady=(0, 5))

        self.btn_admin_login = ctk.CTkButton(self.admin_login_frame, text="Unlock Registration", command=self.verify_admin, height=40, fg_color="#8e44ad", hover_color="#9b59b6")
        self.btn_admin_login.grid(row=3, column=0, pady=(5, 15))

        # 2B. User Details Container (Hidden by default)
        self.user_details_frame = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="#2b2b2b")
        self.user_details_frame.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkLabel(self.user_details_frame, text="Full Name", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=20, pady=(15, 0), sticky="w")
        self.entry_name = ctk.CTkEntry(self.user_details_frame, placeholder_text="Enter full name...", height=40, font=ctk.CTkFont(size=14))
        self.entry_name.grid(row=1, column=0, padx=20, pady=(5, 20), sticky="ew")

        ctk.CTkLabel(self.user_details_frame, text="Roll Number", font=ctk.CTkFont(size=14)).grid(row=0, column=1, padx=20, pady=(15, 0), sticky="w")
        self.entry_roll = ctk.CTkEntry(self.user_details_frame, placeholder_text="Enter roll number...", height=40, font=ctk.CTkFont(size=14))
        self.entry_roll.grid(row=1, column=1, padx=20, pady=(5, 20), sticky="ew")

        # 3. Camera Container
        self.camera_frame = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="#2b2b2b")
        self.camera_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        self.camera_frame.grid_rowconfigure(1, weight=1)
        self.camera_frame.grid_columnconfigure(0, weight=1)

        cam_header = ctk.CTkFrame(self.camera_frame, fg_color="transparent", height=40)
        cam_header.grid(row=0, column=0, sticky="ew", padx=20, pady=10)
        cam_header.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(cam_header, text="Camera Live Feed", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0)
        ctk.CTkLabel(cam_header, text="● LIVE FEED", text_color="#2ecc71", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=1, sticky="e")

        self.video_bg = tk.Frame(self.camera_frame, bg="black")
        self.video_bg.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))
        self.video_label = tk.Label(self.video_bg, bg="black")
        self.video_label.pack(expand=True)

        # 4. Bottom Action Bar
        self.action_bar = ctk.CTkFrame(self.main_frame, height=80, corner_radius=10, fg_color="#2b2b2b")
        self.action_bar.grid(row=3, column=0, sticky="ew")
        self.action_bar.grid_columnconfigure(0, weight=1)
        
        # PREVENT THE BAR FROM COLLAPSING TO FIX JUMPY LAYOUT
        self.action_bar.grid_propagate(False) 

        self.status_var = tk.StringVar(value="System Ready. Please select an option from the menu.")
        self.status_label = ctk.CTkLabel(self.action_bar, textvariable=self.status_var, font=ctk.CTkFont(size=15))
        self.status_label.grid(row=0, column=0, padx=20, pady=20, sticky="w")

        self.action_btn = ctk.CTkButton(self.action_bar, text="Action", height=45, width=160, font=ctk.CTkFont(size=15, weight="bold"))
        self.action_btn.grid(row=0, column=1, padx=20, pady=20)
        self.action_btn.grid_remove()

    # ==================== CAMERA LOGIC ====================
    def start_camera(self):
        if not self.is_camera_running:
            self.cap = cv2.VideoCapture(0)
            self.is_camera_running = True
            self.status_var.set("Camera Active. System Ready.")
            self.update_video_stream()

    def update_video_stream(self):
        if self.is_camera_running:
            ret, frame = self.cap.read()
            if ret:
                frame = cv2.flip(frame, 1) 
                self.current_frame = frame.copy()
                
                h, w, _ = frame.shape
                target_width = 800 
                target_height = int(h * (target_width / w))
                
                frame = cv2.resize(frame, (target_width, target_height))
                cv_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(cv_img)
                imgtk = ImageTk.PhotoImage(image=img)
                
                self.video_label.imgtk = imgtk
                self.video_label.configure(image=imgtk)
            
            self.root.after(15, self.update_video_stream)

    # ==================== ADMIN & REGISTRATION LOGIC ====================
    def verify_admin(self):
        entered_pass = self.entry_admin_pass.get()
        if entered_pass == self.admin_password:
            self.is_admin_authenticated = True
            self.entry_admin_pass.delete(0, 'end')
            self.lbl_login_error.configure(text="")
            self.admin_login_frame.grid_remove()
            self.register_ui() 
        else:
            self.lbl_login_error.configure(text="❌ Incorrect Admin Password")
            self.entry_admin_pass.delete(0, 'end')

    def register_ui(self):
        self.user_details_frame.grid_remove()
        self.action_btn.grid_remove()

        if not self.is_admin_authenticated:
            self.title_label.configure(text="Authentication Required")
            self.status_var.set("Admin rights are required to register new users.")
            self.admin_login_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
            self.lbl_login_error.configure(text="") 
            return

        self.title_label.configure(text="Registering New User")
        self.status_var.set("Position your face, look at the camera, and click 'Capture Face'.")
        
        self.user_details_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        
        self.action_btn.configure(state="normal", text="📷 Capture Face", fg_color="#e67e22", hover_color="#d35400", command=self.process_registration)
        self.action_btn.grid()

    def process_registration(self):
        name = self.entry_name.get().strip()
        roll = self.entry_roll.get().strip()

        if not name or not roll:
            messagebox.showwarning("Missing Information", "Please enter both Full Name and Roll Number.")
            return

        if self.current_frame is None: 
            return

        self.reg_name = name
        self.reg_roll = roll

        filename = f"{self.reg_name}_{self.reg_roll}.jpg"
        filepath = os.path.join(self.db_folder, filename)
        
        cv2.imwrite(filepath, self.current_frame)
        self.status_var.set(f"✅ Successfully registered: {self.reg_name}. Admin session closed.")
        
        self.entry_name.delete(0, 'end')
        self.entry_roll.delete(0, 'end')
        self.user_details_frame.grid_remove() 
        self.action_btn.grid_remove()
        self.title_label.configure(text="System Dashboard")

        self.is_admin_authenticated = False

        for f in os.listdir(self.db_folder):
            if f.endswith(".pkl"):
                os.remove(os.path.join(self.db_folder, f))

    # ==================== ATTENDANCE LOGIC ====================
    def attendance_ui(self):
        self.user_details_frame.grid_remove()
        self.admin_login_frame.grid_remove()

        valid_files = [f for f in os.listdir(self.db_folder) if f.endswith(('.jpg', '.jpeg', '.png'))]
        if not valid_files:
            messagebox.showwarning("Warning", "Database is empty! Please register a user first.")
            return

        self.title_label.configure(text="Mark Attendance")
        self.status_var.set("Look at the camera and click 'Scan Face'.")
        
        # Ensure button is active and visible
        self.action_btn.configure(state="normal", text="🔍 Scan Face", fg_color="#2ecc71", hover_color="#27ae60", command=self.trigger_verification)
        self.action_btn.grid()

    def trigger_verification(self):
        if self.current_frame is None: return
        
        frame_to_check = self.current_frame.copy()
        
        # Disable button instead of removing it to fix layout jumping
        self.action_btn.configure(state="disabled", text="⏳ Processing...", fg_color="#7f8c8d")
        self.status_var.set("Processing Fast AI Recognition...")
        
        threading.Thread(target=self.verify_face_thread, args=(frame_to_check,), daemon=True).start()

    def verify_face_thread(self, frame):
        try:
            dfs = DeepFace.find(
                img_path=frame,
                db_path=self.db_folder,
                model_name="Facenet",
                enforce_detection=False,
                detector_backend="opencv",
                silent=True
            )

            if len(dfs) > 0 and not dfs[0].empty:
                matched_path = dfs[0].iloc[0]['identity']
                filename = os.path.basename(matched_path)
                data = os.path.splitext(filename)[0]
                name, roll = data.split("_")
                
                self.root.after(0, self.mark_attendance_record, name, roll)
            else:
                self.root.after(0, self.status_var.set, "❌ Face not recognized. Try again.")
                # Turn button back on if failed
                self.root.after(0, lambda: self.action_btn.configure(state="normal", text="🔍 Scan Face", fg_color="#2ecc71", hover_color="#27ae60"))

        except Exception as e:
            self.root.after(0, self.status_var.set, f"⚠️ Error during verification.")
            # Turn button back on if failed
            self.root.after(0, lambda: self.action_btn.configure(state="normal", text="🔍 Scan Face", fg_color="#2ecc71", hover_color="#27ae60"))

    def mark_attendance_record(self, name, roll):
        wb = load_workbook(self.excel_file)
        ws = wb.active

        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")

        already_marked = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == name and str(row[1]) == str(roll) and row[2] == date:
                already_marked = True
                break

        if not already_marked:
            ws.append([name, roll, date, time])
            wb.save(self.excel_file)
            success_text = f"✅ Attendance Marked: {name} ({roll})"
            self.status_var.set(success_text)
            messagebox.showinfo("Success", success_text)
        else:
            info_text = f"ℹ️ {name} is already marked for today."
            self.status_var.set(info_text)
            messagebox.showinfo("Info", info_text)
        
        # Re-enable the button once processing is finished
        self.action_btn.configure(state="normal", text="🔍 Scan Face", fg_color="#2ecc71", hover_color="#27ae60", command=self.trigger_verification)
        
        self.root.after(3000, lambda: self.status_var.set("Look at the camera and click 'Scan Face'."))

    def on_closing(self):
        self.is_camera_running = False 
        if self.cap:
            self.cap.release()
        self.root.destroy()

# ==================== RUN APPLICATION ====================
if __name__ == "__main__":
    app_root = ctk.CTk()
    app = FaceAttendanceApp(app_root)
    app_root.mainloop()